# -*- coding: utf-8 -*-
"""
Created on Fri May  8 15:20:20 2020

@author: Brian
Loads indiviudal excel files of AMS data and prepares headers
CO2, CH4, CO in mg/u3
Counts gaps in data
"""



import glob
import pandas as pd
import numpy as np
from openpyxl import Workbook

def clean_name(text_list):
    new_list = []
    for text in text_list:
        new_text = text.replace('Conc,','')
        new_text = new_text.replace(',1','')
        new_text = new_text.replace(',2','')
        new_text = new_text.replace(' conc','')
        new_text = new_text.replace(' Conc','')
        new_list.append(new_text)
    return new_list

file_list = glob.glob('*.xlsx')

#### prepare output Excel file
file_out = 'Kuwait_AMS_GAPS.xls'
workbook = Workbook()  ### make new Excel workbook


for file in file_list:    
    print('\nReading file: ' + file)
        
    df = pd.read_excel(file)
    
    # make new worksheet
    work_sheet = workbook.create_sheet(file[0:15])

    a_headers = clean_name(df.iloc[3,:].to_list()) # extract header names and remove extra text
    df.columns = a_headers  # replace header names
    df.drop(df.index[[0,1,2,3,4,5]], inplace=True) # delete first 6 rows
    df = df.reset_index(drop=True)
    
    df_datetime = df[['Date']].apply(pd.to_datetime)
    #df_datetime['Date']=df_datetime['Date'].dt.date
    #df_datetime['Time']=df_datetime['Time'].dt.time
    
    parameters = a_headers[2:len(a_headers)] #['O3', 'SO2', 'NO2', 'CO', 'PM10', 'PM2.5', 'Wind Direction',
               #'Wind Speed', 'Temperature', 'Relative Humidity', 'Pressure']
    df_aqi = df[parameters].apply(pd.to_numeric)
    df_aqi.CO = df_aqi.CO / 1000 # convert to mg/m3
    
    ########################## begin looking for gaps
    df_null = pd.notnull(df_aqi) + 0 # find nulls (0 is nan, 1 is a number)
    
    col_count = 1
    for analyte in parameters:
        
        print('Finding gaps in: ' + analyte, end = '')
        df_col = df_null[analyte]
        df_col = df_col.reset_index()  # add index column
        df_col = df_col.loc[(df_col[analyte]!=0)] # drop rows with no value (0)
        df_col = df_col.reset_index(drop=True)  # reset index to 0
        
        x_col = df_col['index'].values
        x_col_shift = np.roll(x_col,1)
        
        x_delta = x_col - x_col_shift
        x_delta = x_delta[1:len(x_delta)-1] # drop first and last elements from shift
        
        x_delta_merge = x_delta - 1  # convert to 0
        
        gaps = x_delta_merge[x_delta_merge != 0]  # remove all 0 rows
        gaps = gaps[gaps < 1400]  # remove all rows > 1400
        
        print(' ---- ' + str(len(gaps)) + ' gaps found.')
        
        work_sheet.cell(row = 1, column = col_count).value = analyte # make column name
        for i in range(1,len(gaps)):
            work_sheet.cell(row = i+1, column = col_count).value = gaps[i]
        
        col_count += 1 # increment counter
        
    print('\nSaving in file: ' + file_out)
    workbook.save(filename=file_out)




    